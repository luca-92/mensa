import csv
from django.http import HttpResponse
from django.shortcuts import render
import datetime
from openpyxl import Workbook
import pandas as pd
from django.core.files.storage import FileSystemStorage

from prenotamensa.models import Personale, AssociazioneAtt

# Create your views here.
def index(request):

    now = datetime.datetime.now()
    breakfast_time_start = now.replace( hour=6, minute=50, second=0, microsecond=0 )
    breakfast_time_finish = now.replace( hour=8, minute=10, second=0, microsecond=0 )

    lunch_time_start = now.replace( hour=11, minute=20, second=0, microsecond=0 )
    lunch_time_finish = now.replace( hour=13, minute=10, second=0, microsecond=0 )

    dinner_time_start = now.replace( hour=17, minute=20, second=0, microsecond=0 )
    dinner_time_finish = now.replace( hour=19, minute=10, second=0, microsecond=0 )

    sandwich_time_start = now.replace( hour=20, minute=50, second=0, microsecond=0 )
    sandwich_time_finish = now.replace( hour=23, minute=59, second=0, microsecond=0 )

    to_disable = {"colazione":"","pranzo":"","cena":"","sandwich":""}
    if (breakfast_time_start <= now <= breakfast_time_finish):
        to_disable["pranzo"]="disable"
        to_disable["cena"]="disable"
        to_disable["sandwich"]="disable"
    elif (lunch_time_start <= now <= lunch_time_finish):
        to_disable["colazione"]="disable"
        to_disable["cena"]="disable"
        to_disable["sandwich"]="disable"
    elif (dinner_time_start <= now <= dinner_time_finish):
        to_disable["colazione"]="disable"
        to_disable["pranzo"]="disable"
        to_disable["sandwich"]="disable"
    elif (sandwich_time_start <= now <= sandwich_time_finish):
        to_disable["colazione"]="disable"
        to_disable["pranzo"]="disable"
        to_disable["cena"]="disable"
    else:
        to_disable["colazione"]="disable"
        to_disable["pranzo"]="disable"
        to_disable["cena"]="disable"
        to_disable["sandwich"]="disable"

    return render(request, 'prenotamensa/index.html', {"to_disable":to_disable})

def check(request):
    numeroCMD = request.POST.get("numerocmd", "").strip().upper()
    pasto = request.POST.get('optradio',"")
    lavanderia = request.POST.get('optradiolav',"")
    print("lavanderia: " + lavanderia)
    try:
        persona = Personale.objects.get(numero_CMD = numeroCMD)
    except:
        return render(request, 'prenotamensa/errore.html', {"errore":"ERRORE: CMD INESISTENTE","show":"true"})
    to_compare = AssociazioneAtt.objects.filter(numero_CMD = numeroCMD, datapasto = datetime.date.today())
    if not to_compare:
        if pasto == "colazione" and lavanderia == "lavanderia":
           AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="X", Pranzo="", Cena="", Sandwich="", Lavanderia="X", datapasto=datetime.datetime.today() )
        elif pasto == "pranzo" and lavanderia == "lavanderia":
            AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="", Pranzo="X", Cena="", Sandwich="", Lavanderia="X", datapasto=datetime.datetime.today() )
        elif pasto == "cena" and lavanderia == "lavanderia":
            AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="", Pranzo="", Cena="X", Sandwich="", Lavanderia="X",datapasto=datetime.datetime.today() )
        elif pasto == "sandwich" and lavanderia == "lavanderia":
            AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="", Pranzo="", Cena="", Sandwich="X", Lavanderia="X",datapasto=datetime.datetime.today() )
        elif pasto == "colazione" and lavanderia == "":
            AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="X", Pranzo="", Cena="", Sandwich="", Lavanderia="",datapasto=datetime.datetime.today() )
        elif pasto == "pranzo" and lavanderia == "":
            AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="", Pranzo="X", Cena="", Sandwich="", Lavanderia="",datapasto=datetime.datetime.today() )
        elif pasto == "cena" and lavanderia == "":
            AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="", Pranzo="", Cena="X", Sandwich="", Lavanderia="",datapasto=datetime.datetime.today() )
        elif pasto == "sandwich" and lavanderia == "":
            AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="", Pranzo="", Cena="", Sandwich="X", Lavanderia="",datapasto=datetime.datetime.today() )
        elif lavanderia == "lavanderia" and pasto == "":
            AssociazioneAtt.objects.create(numero_CMD = numeroCMD, Colazione="", Pranzo="", Cena="", Sandwich="", Lavanderia="X",datapasto=datetime.datetime.today() )
    else:
        to_update = to_compare[0]
        if pasto == "colazione":
            if to_update.Colazione == "X":
                return render(request, 'prenotamensa/errore.html', {"errore":"ERRORE: " + persona.nominativo + " OGGI HAI GIA' PRENOTATO LA COLAZIONE"})
            to_update.Colazione = "X"
        elif pasto == "pranzo":
            if to_update.Pranzo == "X":
                return render(request, 'prenotamensa/errore.html', {"errore":"ERRORE: " + persona.nominativo + " OGGI HAI GIA' PRENOTATO IL PRANZO"})
            to_update.Pranzo = "X"
        elif pasto == "cena":
            if to_update.Cena == "X":
                return render(request, 'prenotamensa/errore.html', {"errore":"ERRORE: " + persona.nominativo + " OGGI HAI GIA' PRENOTATO LA CENA"})
            to_update.Cena = "X"
        elif pasto == "sandwich":
            if to_update.Sandwich == "X":
                return render(request, 'prenotamensa/errore.html', {"errore":"ERRORE: " + persona.nominativo + " OGGI HAI GIA' PRENOTATO IL SANDWICH"})
            to_update.Sandwich = "X"
        if lavanderia == "lavanderia":
            if to_update.Lavanderia == "X":
                return render(request, 'prenotamensa/errore.html', {"errore":"ERRORE: " + persona.nominativo + " OGGI HAI GIA' PRENOTATO LA LAVANDERIA"})
            to_update.Lavanderia = "X"
        to_update.save()
    messaggio1 = "Ciao " + str(persona.nominativo) + ". "
    messaggio2 = ""
    if pasto != "" and lavanderia == "lavanderia":
        messaggio2 += "Prenotazione " + pasto + " e lavanderia completata"
    elif pasto == "" and lavanderia == 'lavanderia':
        messaggio2 += "Prenotazione lavanderia completata"
    else:
        messaggio2 += "Prenotazione " + pasto + " completata"
    return render(request, 'prenotamensa/ok.html',{'messaggio1': messaggio1, "messaggio2":messaggio2})

def excel(request):

    if request.method == 'POST':
        giorno = request.POST.get("esporta", "")
        tot_colazione = 0
        tot_pranzo = 0
        tot_cena = 0
        tot_sandwich = 0
        tot_lavanderia = 0
        totale = 0
        lista = AssociazioneAtt.objects.filter(datapasto = giorno)
        for personale in lista:
            print(personale.numero_CMD)
            persona = Personale.objects.get(numero_CMD = personale.numero_CMD)
            personale.numero_CMD = persona.nominativo
            if personale.Colazione == "X":
                tot_colazione += 1
            if personale.Pranzo == "X":
                tot_pranzo += 1
            if personale.Cena == "X":
                tot_cena += 1
            if personale.Sandwich == "X":
                tot_sandwich += 1
            if personale.Lavanderia == "X":
                tot_lavanderia += 1
            totale +=1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Associazioni.xlsx"'        



        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = giorno
        cell = worksheet.cell(1,1)
        cell.value = giorno

        columns = [
            'N°',
            'Nominativo',
            'Colazione',
            'Pranzo',
            'Cena',
            'Sandwich',
            'Lavanderia',
        ]

        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
        indice = 1
        for associazione in lista:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [indice, associazione.numero_CMD,associazione.Colazione, associazione.Pranzo, associazione.Cena,associazione.Sandwich,associazione.Lavanderia]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
            indice +=1

        row_num += 1
        cell = worksheet.cell(row_num,1)
        cell.value = "TOTALE"

        cell = worksheet.cell(row_num,2)
        cell.value = totale

        cell = worksheet.cell(row_num,3)
        cell.value = tot_colazione

        cell = worksheet.cell(row_num,4)
        cell.value = tot_pranzo

        cell = worksheet.cell(row_num,5)
        cell.value = tot_cena

        cell = worksheet.cell(row_num,6)
        cell.value = tot_sandwich

        cell = worksheet.cell(row_num,7)
        cell.value = tot_lavanderia

        workbook.save(response)

        return response
 
    return render(request, 'prenotamensa/excel.html')

def excel2(request):
    if request.method == 'POST':
        mese = request.POST.get("dropdown", "")
        max_giorni = 32
        if str(mese) == "02":
            max_giorni = 29
        elif (str(mese) == "04") or (str(mese) == "06") or (str(mese) == "09") or (str(mese) == "11"):
            max_giorni = 31
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="'+str(mese)+'2023.xlsx"'        

        workbook = Workbook()

        for i in range(1,max_giorni):
            tot_colazione = 0
            tot_pranzo = 0
            tot_cena = 0
            tot_sandwich = 0
            tot_lavanderia = 0
            totale = 0
            giorno = "2023-"+str(mese)+"-"+str(i)
            lista = AssociazioneAtt.objects.filter(datapasto = giorno)
            for personale in lista:
                print(personale.numero_CMD)
                persona = Personale.objects.get(numero_CMD = personale.numero_CMD)
                personale.numero_CMD = persona.nominativo
                if personale.Colazione == "X":
                    tot_colazione += 1
                if personale.Pranzo == "X":
                    tot_pranzo += 1
                if personale.Cena == "X":
                    tot_cena += 1
                if personale.Sandwich == "X":
                    tot_sandwich += 1
                if personale.Lavanderia == "X":
                    tot_lavanderia += 1
                totale +=1
            
            worksheet = workbook.create_sheet(str(i))
            cell = worksheet.cell(1,1)
            cell.value = giorno

            columns = [
                'N°',
                'Nominativo',
                'Colazione',
                'Pranzo',
                'Cena',
                'Sandwich',
                'Lavanderia',
            ]

            row_num = 2

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
            indice = 1
            for associazione in lista:
                row_num += 1
                
                # Define the data for each cell in the row 
                row = [indice, associazione.numero_CMD,associazione.Colazione, associazione.Pranzo, associazione.Cena,associazione.Sandwich,associazione.Lavanderia]
                
                # Assign the data for each cell of the row 
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                indice +=1

            row_num += 1
            cell = worksheet.cell(row_num,1)
            cell.value = "TOTALE"

            cell = worksheet.cell(row_num,2)
            cell.value = totale

            cell = worksheet.cell(row_num,3)
            cell.value = tot_colazione

            cell = worksheet.cell(row_num,4)
            cell.value = tot_pranzo

            cell = worksheet.cell(row_num,5)
            cell.value = tot_cena

            cell = worksheet.cell(row_num,6)
            cell.value = tot_sandwich

            cell = worksheet.cell(row_num,7)
            cell.value = tot_lavanderia
        workbook.remove(workbook['Sheet'])
        workbook.save(response)

        return response
 
    return render(request, 'prenotamensa/excel.html')

def aggiorna(request):
    df = pd.read_excel('C:\\Users\\User\\Downloads\\CMD.xlsx', index_col=0)
    new_pax = 0
    for index, row in df.iterrows():
        try:
            Personale.objects.get(numero_CMD = row['CMD'])
        except:
            new_pax += 1
            nominativo = str(row['grado']) + " " + str(row['nome']) + " " + str(row['cognome'])
            Personale.objects.create(numero_CMD = row['CMD'], nominativo = nominativo)

    messaggio2 = "Operazione Completata, aggiunte " + str(new_pax) + " PAX."
    return render(request, 'prenotamensa/ok.html',{'messaggio1': "", "messaggio2":messaggio2})

def insert(request):
    if request.method == 'POST':
        nominativo = request.POST.get("nominativo", "")
        numerocmd = request.POST.get("numerocmd", "").strip().upper()
        try:
            Personale.objects.get(numero_CMD = numerocmd)
        except:
            Personale.objects.create(numero_CMD = numerocmd, nominativo = nominativo)
    messaggio2 = "Operazione Completata, aggiunte 1 PAX."
    return render(request, 'prenotamensa/ok.html',{'messaggio1': "", "messaggio2":messaggio2})
